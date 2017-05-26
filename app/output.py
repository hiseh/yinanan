from openpyxl import Workbook
from datetime import datetime
from unittest import TestCase
from app.kousuan import KouSuan

__author__ = 'hisehyin'
__datetime__ = '15/12/25 下午4:46'


class Excel:
    @staticmethod
    def grouped(iterable, n):
            return zip(*[iter(iterable)] * n)

    @staticmethod
    def output(file_path, questions):
        """
        输出成excel
        :param file_path:
        :param questions:
        :return:
        """
        wb = Workbook()
        ws = wb.create_sheet(title='questions')
        for i, e in enumerate(Excel.grouped(questions, 5)):
            for c, q in enumerate(e):
                ws['{column}{line}'.format(column=''.join(chr(65 + c)), line=i + 1)] = q

        return wb.save(filename=file_path)


class Table:
    @staticmethod
    def kousuan():
        questions = KouSuan.gen_kousuan_questions(120, 100)
        for e in zip(*[iter(set(questions))] * 5):
            print('\t'.join(e))

    @staticmethod
    def shushi():
        questions = KouSuan.gen_add_sub_questions(total=24, max_=1000, add_carry=True, sub_carry=True)
        for e in zip(*[iter(set(questions))] * 4):
            print('\t'.join(e))


class Test(TestCase):
    def setUp(self):
        self.start = datetime.now()

    def tearDown(self):
        print(datetime.now() - self.start)

    def test_excel_output(self):
        result = Excel.output('/Users/hiseh/temp/an.xlsx', KouSuan.gen_add_sub_questions(1200, 100, 9, 0, True, True))
        print(result)

    def test_table_kousuan(self):
        Table.kousuan()

    def test_shushi(self):
        Table.shushi()
